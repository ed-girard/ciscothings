import os, re
from datetime import datetime
from openpyxl import load_workbook
from jinja2 import Template, Environment, FileSystemLoader
j2_env = Environment(loader=FileSystemLoader('TEMPLATES'),trim_blocks=False)
date_string = f'{datetime.now():%Y-%m-%d}'

def ap_mac(mac_str):
    bad_chars = re.compile('\.|:')
    mac_str = bad_chars.sub('',mac_str)
    return (mac_str[0:4]+"."+mac_str[4:8]+"."+mac_str[8:12])

#Create apDict for kinja template
apDict = {}

#Load Workbook
wb = load_workbook(filename = 'ap_import.xlsx')

#Select Config to get SITECODE and GROUP names
sheet = wb['Config']
apDict['SITECODE'] = sheet.cell(row=1,column=2).value
apDict['GROUP'] = sheet.cell(row=2,column=2).value

#Select APList to begin config loop
sheet = wb['APList']
max_row = sheet.max_row
max_column = sheet.max_column

#Delete prior output for cleanup
try:
    os.remove(date_string+"_"+apDict['GROUP']+"_wlc_cli.txt")
except OSError:
    pass

#For each row get MAC and hostname, put into dict, use template to append to output
for i in range (2,max_row+1):
    apDict['MAC'] = ap_mac(sheet.cell(row=i,column=1).value.upper())
    apDict['NAME'] = sheet.cell(row=i,column=2).value.upper()
    template = j2_env.get_template('assa_add_access_point.jinja')
    f = open(date_string+"_"+apDict['GROUP']+"_wlc_cli.txt",'a+')
    f.write(template.render(apDict=apDict))

f.close